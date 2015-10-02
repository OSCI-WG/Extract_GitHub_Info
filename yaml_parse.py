__author__ = 'Majisto'
import sys
import os
import tkinter as tk
from tkinter import filedialog
import yaml
from openpyxl import Workbook

class Member(yaml.YAMLObject):
    yaml_tag = "Member"
    def __init__(self, login, id_number, full_name, two_factor, email):
        self.login = login
        self.full_name = full_name
        self.user_email = email
        self.number = id_number
        self.two_factor = two_factor

def GetWorkBook() :
    wb = Workbook()
    ws = wb.active
    ws.title = "Organization Member Information"
    ws['A1'] = 'GitHub ID'
    ws['B1'] = 'Full Name'
    ws['C1'] = 'Email'
    ws['D1'] = 'Two-Factor'
    return wb, ws

def ProcessInputFile(member_list):
    wb, ws = GetWorkBook()
    i = 2
    for member in member_list:
        assert isinstance(member, Member)
        ws['A%d' % (i,)] = member.login
        ws['B%d' % (i,)] = member.full_name
        ws['C%d' % (i,)] = member.user_email
        ws['D%d' % (i,)] = "Yes" if member.two_factor else "No"
        i+=1
    SaveWorkbook(wb)

def SaveWorkbook(wb):
    filename = input("Please enter a filename to save as: ")
    if not filename.endswith('.xlsx'):
        filename += ".xlsx"
    while True:
        if os.path.isfile(filename): #Workbook already exists.  Confirm over-write.
            if input("File already exists.  Are you sure you want to save? ") in ('y' , 'Y') :
                wb.save(filename)
                return
            else:
                filename = input("Please enter a new filename without extension: ")
                if not filename.endswith('.xlsx'):
                    filename += ".xlsx"
        else:
            wb.save(filename)
            return

def GetYamlFile():
    for x in sys.argv:
        if x.endswith(".yaml"):
            return open (x, "r")
    root = tk.Tk()
    root.withdraw()
    options = {"filetypes": [('all_files', '.yaml')]}
    file_path = tk.filedialog.askopenfilename(**options)
    try:
        return open (file_path, "r")
    except FileNotFoundError:
        print("Unable to find file specified.")
        sys.exit(-1)

def main():
    input_file = GetYamlFile()
    member_list = yaml.load(input_file)
    ProcessInputFile(member_list)

main()